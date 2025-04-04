from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, FormView, View
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
from django.template.loader import render_to_string, get_template
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from xhtml2pdf import pisa
import pytz
import io

from .models import Lead, LeadSource, LeadStatus
from .forms import LeadForm, LeadUpdateForm, LeadAssignForm
from clients.forms import ClientForm
from accounts.models import CustomUser

class AdminOrManagerRequiredMixin(UserPassesTestMixin):
    """Verify that the current user is admin or manager"""
    def test_func(self):
        return self.request.user.is_authenticated and (
            self.request.user.is_superuser or 
            self.request.user.role in ['admin', 'manager']
        )

class LeadAccessMixin(UserPassesTestMixin):
    """Verify that the current user has access to the lead"""
    def test_func(self):
        if not self.request.user.is_authenticated:
            return False
            
        # Admin and managers can access all leads
        if self.request.user.is_superuser or self.request.user.role in ['admin', 'manager']:
            return True
            
        # Get the lead object
        lead_id = self.kwargs.get('pk')
        lead = get_object_or_404(Lead, pk=lead_id)
        
        # Team leaders can access their own and their team members' leads
        if self.request.user.role == 'team_leader':
            team_members = CustomUser.objects.filter(
                Q(role='sales_rep') & 
                (Q(profile__team_leader=self.request.user) | Q(pk=self.request.user.pk))
            )
            return lead.assigned_to in team_members
            
        # Sales reps can only access their own leads
        return lead.assigned_to == self.request.user

class LeadListView(LoginRequiredMixin, ListView):
    """Display all leads"""
    model = Lead
    template_name = 'leads/lead_list.html'
    context_object_name = 'leads'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = Lead.objects.all()
        
        # # Filter by role
        # if self.request.user.role == 'sales_rep':
        #     queryset = queryset.filter(assigned_to=self.request.user)
        # elif self.request.user.role == 'team_leader':
        #     team_members = CustomUser.objects.filter(
        #         Q(role='sales_rep') & 
        #         (Q(profile__team_leader=self.request.user) | Q(pk=self.request.user.pk))
        #     )
        #     queryset = queryset.filter(assigned_to__in=team_members)
            
        # # Search and filter options
        # search_query = self.request.GET.get('search', '')
        # status_filter = self.request.GET.get('status', '')
        
        # if search_query:
        #     queryset = queryset.filter(
        #         Q(name__icontains=search_query) |
        #         Q(company__icontains=search_query) |
        #         Q(email__icontains=search_query) |
        #         Q(phone__icontains=search_query)
        #     )
            
        # if status_filter:
        #     queryset = queryset.filter(status__id=status_filter)
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['statuses'] = LeadStatus.objects.all()
        context['active_status'] = self.request.GET.get('status', '')
        context['search_query'] = self.request.GET.get('search', '')
        return context

class LeadDetailView(LoginRequiredMixin, LeadAccessMixin, DetailView):
    """Display lead details"""
    model = Lead
    template_name = 'leads/lead_detail.html'
    context_object_name = 'lead'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        lead = self.get_object()
        
        # If the lead is converted, add client form
        if lead.is_converted:
            context['client_form'] = ClientForm(initial={
                'name': lead.name,
                'company': lead.company,
                'phone': lead.phone,
                'email': lead.email,
                'address': lead.address,
                'lead': lead,
            })
            
        return context

class LeadCreateView(LoginRequiredMixin, CreateView):
    """Create a new lead"""
    model = Lead
    form_class = LeadForm
    template_name = 'leads/lead_form.html'
    success_url = reverse_lazy('lead_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        
        # If the user is a sales rep or team leader, auto-assign the lead to them
        if self.request.user.role in ['sales_rep', 'team_leader']:
            form.instance.assigned_to = self.request.user
        
        messages.success(self.request, 'Lead created successfully.')
        return super().form_valid(form)

class LeadUpdateView(LoginRequiredMixin, LeadAccessMixin, UpdateView):
    """Update an existing lead"""
    model = Lead
    form_class = LeadUpdateForm
    template_name = 'leads/lead_form.html'
    
    def form_valid(self, form):
        messages.success(self.request, 'Lead updated successfully.')
        return super().form_valid(form)

class LeadDeleteView(LoginRequiredMixin, AdminOrManagerRequiredMixin, DeleteView):
    """Delete a lead"""
    model = Lead
    template_name = 'leads/lead_confirm_delete.html'
    success_url = reverse_lazy('lead_list')
    
    def delete(self, request, *args, **kwargs):
        lead = self.get_object()
        messages.success(request, f'Lead "{lead}" deleted successfully.')
        return super().delete(request, *args, **kwargs)

class LeadAssignView(LoginRequiredMixin, AdminOrManagerRequiredMixin, UpdateView):
    """Assign a lead to a user"""
    model = Lead
    form_class = LeadAssignForm
    template_name = 'leads/lead_assign.html'
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        
        # Managers can only assign to team leaders and sales reps
        if self.request.user.role == 'manager':
            form.fields['assigned_to'].queryset = CustomUser.objects.filter(
                role__in=['team_leader', 'sales_rep']
            )
            
        return form
    
    def form_valid(self, form):
        lead = self.get_object()
        user = form.cleaned_data['assigned_to']
        messages.success(self.request, f'Lead assigned to {user.get_full_name() or user.username} successfully.')
        return super().form_valid(form)

class MyLeadsView(LoginRequiredMixin, ListView):
    """Display leads assigned to the current user"""
    model = Lead
    template_name = 'leads/lead_list.html'
    context_object_name = 'leads'
    paginate_by = 10
    
    def get_queryset(self):
        return Lead.objects.filter(assigned_to=self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['statuses'] = LeadStatus.objects.all()
        context['is_my_leads'] = True
        return context
        
class ExportLeadsExcelView(LoginRequiredMixin, View):
    """Export leads to Excel"""
    def get(self, request):
        # Get leads based on user role and filters (similar to LeadListView)
        queryset = Lead.objects.all()
        
        # Filter by role
        if request.user.role == 'sales_rep':
            queryset = queryset.filter(assigned_to=request.user)
        elif request.user.role == 'team_leader':
            team_members = CustomUser.objects.filter(
                Q(role='sales_rep') & 
                (Q(profile__team_leader=request.user) | Q(pk=request.user.pk))
            )
            queryset = queryset.filter(assigned_to__in=team_members)
            
        # Search and filter options
        search_query = request.GET.get('search', '')
        status_filter = request.GET.get('status', '')
        
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(company__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(phone__icontains=search_query)
            )
            
        if status_filter:
            queryset = queryset.filter(status__id=status_filter)
        
        # Create workbook and add worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        
        # Header styles
        header_font = Font(name='Arial', bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        centered = Alignment(horizontal='center')
        
        # Add headers
        headers = ['ID', 'Name', 'Company', 'Email', 'Phone', 'Status', 'Source', 'Assigned To']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
        
        # Add data
        row_num = 2
        for lead in queryset:
            ws.cell(row=row_num, column=1, value=lead.id)
            ws.cell(row=row_num, column=2, value=lead.name)
            ws.cell(row=row_num, column=3, value=lead.company)
            ws.cell(row=row_num, column=4, value=lead.email)
            ws.cell(row=row_num, column=5, value=lead.phone)
            ws.cell(row=row_num, column=6, value=lead.status.name if lead.status else 'N/A')
            ws.cell(row=row_num, column=7, value=lead.source.name if lead.source else 'N/A')
            ws.cell(row=row_num, column=8, value=str(lead.assigned_to) if lead.assigned_to else 'Unassigned')
            row_num += 1
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=leads_export.xlsx'
        
        # Save workbook to response
        wb.save(response)
        return response

class ExportLeadsPDFView(LoginRequiredMixin, View):
    """Export leads to PDF"""
    def get(self, request):
        # Get leads based on user role and filters (similar to LeadListView)
        queryset = Lead.objects.all()
        
        # Filter by role
        if request.user.role == 'sales_rep':
            queryset = queryset.filter(assigned_to=request.user)
        elif request.user.role == 'team_leader':
            team_members = CustomUser.objects.filter(
                Q(role='sales_rep') & 
                (Q(profile__team_leader=request.user) | Q(pk=request.user.pk))
            )
            queryset = queryset.filter(assigned_to__in=team_members)
            
        # Search and filter options
        search_query = request.GET.get('search', '')
        status_filter = request.GET.get('status', '')
        
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(company__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(phone__icontains=search_query)
            )
            
        if status_filter:
            queryset = queryset.filter(status__id=status_filter)
        
        # Render template to string
        html_string = render_to_string(
            'leads/lead_pdf_template.html',
            {
                'leads': queryset,
                'title': 'Leads Report',
                'date': timezone.now().strftime('%B %d, %Y'),
                'user': request.user,
            }
        )
        
        # Create PDF
        result = io.BytesIO()
        pdf = pisa.pisaDocument(io.BytesIO(html_string.encode("UTF-8")), result)
        
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="leads_report.pdf"'
            return response
        
        return HttpResponse("Error generating PDF", status=500)


import pandas as pd
from django.shortcuts import render
from django.contrib import messages
from .forms import ExcelUploadForm

def upload_leads(request):
    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES["file"]

            try:
                df = pd.read_excel(file)  # Read Excel File
                
                leads_data = []
                for _, row in df.iterrows():
                    # Check if Source exists, else create it
                    source_name = row.get("Source", "").strip()
                    if source_name:
                        source, _ = LeadSource.objects.get_or_create(name=source_name)
                    else:
                        source = None  # Allow blank sources

                    # Check if Status exists, else create it
                    status_name = row.get("Status", "").strip()
                    if status_name:
                        status, _ = LeadStatus.objects.get_or_create(name=status_name)
                    else:
                        status = None  # Allow blank statuses

                    # Check if Assigned User exists

                    # Append processed data
                    leads_data.append({
                        "name": row.get("Name", ""),
                        "company": row.get("Company", ""),
                        "phone": row.get("Phone", ""),
                        "email": row.get("Email", ""),
                        "address": row.get("Address", ""),
                        "source": source,
                        "status": status,
                        "created_by": request.user,
                    })

                # Bulk Insert Leads
                Lead.objects.bulk_create([Lead(**data) for data in leads_data])

                messages.success(request, "Excel data processed and leads added successfully!")
                return redirect('lead_list')

            except Exception as e:
                messages.error(request, f"Error: {e}")
    
    else:
        form = ExcelUploadForm()

    return render(request, "upload_leads.html", {"form": form})
